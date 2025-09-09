#!/usr/bin/env python3
"""
Spreadsheet Document Extractor Worker

Extracts content from spreadsheet files (XLSX, XLS, ODS, CSV) while preserving
structure, formulas, and formatting information.
"""

import json
import asyncio
from pathlib import Path
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
import hashlib
import csv
import io

import typer
from loguru import logger
from rich.console import Console
from rich.table import Table as RichTable
from rich.progress import Progress, SpinnerColumn, TextColumn

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not available - Excel support limited")

try:
    from odf import opendocument
    from odf.table import Table as OdfTable, TableRow, TableCell
    ODF_AVAILABLE = True
except ImportError:
    ODF_AVAILABLE = False
    logger.warning("odfpy not available - ODS support disabled")

# Configure logger
logger.remove()
logger.add(lambda msg: print(msg, end=""), level="INFO", format="{message}")

app = typer.Typer(help="Extract content from spreadsheet documents")
console = Console()


class SpreadsheetExtractor:
    """Extract content from spreadsheet documents."""
    
    def __init__(self):
        self.cache_dir = Path.home() / ".cache" / "extractor" / "spreadsheet"
        self.cache_dir.mkdir(parents=True, exist_ok=True)
        
    async def extract_spreadsheet(self,
                                 file_path: Path,
                                 sheet_names: Optional[List[str]] = None,
                                 extract_formulas: bool = True,
                                 extract_formatting: bool = True,
                                 max_empty_rows: int = 10) -> Dict:
        """Extract content from spreadsheet file.
        
        Args:
            file_path: Path to spreadsheet file
            sheet_names: Specific sheets to extract (None = all)
            extract_formulas: Whether to extract formulas
            extract_formatting: Whether to extract cell formatting
            max_empty_rows: Stop after this many consecutive empty rows
            
        Returns:
            Extracted content with metadata
        """
        # Validate path
        file_path = Path(file_path).resolve()
        if not file_path.exists():
            raise FileNotFoundError(f"Spreadsheet file not found: {file_path}")
        
        suffix = file_path.suffix.lower()
        if suffix not in ['.xlsx', '.xls', '.xlsm', '.ods', '.csv']:
            raise ValueError(f"Not a spreadsheet file: {file_path}")
        
        # Check cache
        cache_key = self._get_cache_key(file_path)
        cached = await self._check_cache(cache_key)
        if cached:
            logger.info(f"Using cached extraction for {file_path.name}")
            return cached
        
        # Extract based on format
        if suffix == '.csv':
            result = await self._extract_csv(file_path)
        elif suffix in ['.xlsx', '.xls', '.xlsm']:
            if not OPENPYXL_AVAILABLE:
                raise ImportError("openpyxl required for Excel files")
            result = await self._extract_excel(
                file_path, sheet_names, extract_formulas, 
                extract_formatting, max_empty_rows
            )
        elif suffix == '.ods':
            if not ODF_AVAILABLE:
                raise ImportError("odfpy required for ODS files")
            result = await self._extract_ods(
                file_path, sheet_names, max_empty_rows
            )
        else:
            raise ValueError(f"Unsupported format: {suffix}")
        
        # Add file metadata
        result["metadata"]["source_file"] = str(file_path)
        result["metadata"]["extraction_time"] = datetime.now().isoformat()
        result["metadata"]["file_size"] = file_path.stat().st_size
        
        # Cache result
        await self._cache_result(cache_key, result)
        
        return result
    
    async def _extract_csv(self, file_path: Path) -> Dict:
        """Extract content from CSV file."""
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            # Detect delimiter
            sample = f.read(1024)
            f.seek(0)
            
            sniffer = csv.Sniffer()
            try:
                dialect = sniffer.sniff(sample)
            except csv.Error:
                dialect = csv.excel()
            
            reader = csv.reader(f, dialect)
            rows = list(reader)
        
        # Convert to table structure
        cells = []
        headers = []
        
        for row_idx, row in enumerate(rows):
            # First row often headers
            if row_idx == 0 and all(isinstance(cell, str) and cell.strip() for cell in row):
                headers = [0]
            
            for col_idx, value in enumerate(row):
                cells.append({
                    "row": row_idx,
                    "col": col_idx,
                    "value": value,
                    "type": "string"
                })
        
        sheets = [{
            "name": "Sheet1",
            "rows": len(rows),
            "cols": max(len(row) for row in rows) if rows else 0,
            "cells": cells,
            "headers": headers
        }]
        
        return {
            "sheets": sheets,
            "metadata": {
                "format": "csv",
                "delimiter": dialect.delimiter,
                "total_sheets": 1
            },
            "statistics": self._calculate_statistics(sheets)
        }
    
    async def _extract_excel(self, 
                           file_path: Path,
                           sheet_names: Optional[List[str]],
                           extract_formulas: bool,
                           extract_formatting: bool,
                           max_empty_rows: int) -> Dict:
        """Extract content from Excel file."""
        wb = load_workbook(
            file_path, 
            data_only=not extract_formulas,
            keep_vba=False
        )
        
        sheets = []
        
        # Process each worksheet
        for ws in wb.worksheets:
            if sheet_names and ws.title not in sheet_names:
                continue
            
            sheet_data = await self._extract_excel_sheet(
                ws, extract_formulas, extract_formatting, max_empty_rows
            )
            sheets.append(sheet_data)
        
        # Extract named ranges
        named_ranges = {}
        for name, defn in wb.defined_names.items():
            if defn.type == 'RANGE':
                named_ranges[name] = {
                    "range": defn.value,
                    "scope": defn.scope
                }
        
        # Extract document properties
        props = wb.properties
        metadata = {
            "format": "excel",
            "created": props.created.isoformat() if props.created else None,
            "modified": props.modified.isoformat() if props.modified else None,
            "creator": props.creator or "",
            "last_modified_by": props.lastModifiedBy or "",
            "title": props.title or "",
            "total_sheets": len(wb.worksheets),
            "sheet_names": wb.sheetnames
        }
        
        wb.close()
        
        return {
            "sheets": sheets,
            "named_ranges": named_ranges,
            "metadata": metadata,
            "statistics": self._calculate_statistics(sheets)
        }
    
    async def _extract_excel_sheet(self,
                                 worksheet,
                                 extract_formulas: bool,
                                 extract_formatting: bool,
                                 max_empty_rows: int) -> Dict:
        """Extract data from a single Excel worksheet."""
        cells = []
        headers = []
        merged_cells = []
        
        # Track dimensions
        max_row = worksheet.max_row or 0
        max_col = worksheet.max_column or 0
        
        # Detect actual data range
        empty_row_count = 0
        actual_max_row = 0
        
        for row_idx in range(max_row):
            row_empty = True
            row_cells = []
            
            for col_idx in range(max_col):
                cell = worksheet.cell(row=row_idx + 1, column=col_idx + 1)
                
                if cell.value is not None:
                    row_empty = False
                    empty_row_count = 0
                    
                    cell_data = {
                        "row": row_idx,
                        "col": col_idx,
                        "value": str(cell.value) if cell.value is not None else "",
                        "type": self._get_cell_type(cell)
                    }
                    
                    # Extract formula if available
                    if extract_formulas and hasattr(cell, 'formula') and cell.formula:
                        cell_data["formula"] = cell.formula
                    
                    # Extract formatting if requested
                    if extract_formatting:
                        formatting = self._extract_cell_formatting(cell)
                        if formatting:
                            cell_data["formatting"] = formatting
                    
                    row_cells.append(cell_data)
            
            if not row_empty:
                cells.extend(row_cells)
                actual_max_row = row_idx
            else:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    break
        
        # Detect headers (first row with all string values)
        if cells and all(c["type"] == "string" for c in cells if c["row"] == 0):
            headers = [0]
        
        # Extract merged cells
        for merged_range in worksheet.merged_cells.ranges:
            merged_cells.append({
                "start_row": merged_range.min_row - 1,
                "start_col": merged_range.min_col - 1,
                "end_row": merged_range.max_row - 1,
                "end_col": merged_range.max_col - 1
            })
        
        # Extract charts
        charts = []
        if hasattr(worksheet, '_charts'):
            for chart in worksheet._charts:
                charts.append({
                    "type": chart.__class__.__name__,
                    "title": getattr(chart, 'title', '')
                })
        
        return {
            "name": worksheet.title,
            "rows": actual_max_row + 1,
            "cols": max_col,
            "cells": cells,
            "headers": headers,
            "merged_cells": merged_cells,
            "charts": charts,
            "hidden": worksheet.sheet_state == 'hidden'
        }
    
    def _get_cell_type(self, cell) -> str:
        """Determine cell data type."""
        if cell.value is None:
            return "empty"
        elif cell.data_type == 'n':
            return "number"
        elif cell.data_type == 'd':
            return "date"
        elif cell.data_type == 'b':
            return "boolean"
        elif cell.data_type == 'f':
            return "formula"
        elif cell.data_type == 'e':
            return "error"
        else:
            return "string"
    
    def _extract_cell_formatting(self, cell) -> Optional[Dict]:
        """Extract cell formatting information."""
        formatting = {}
        
        # Font formatting
        if cell.font:
            font_info = {}
            if cell.font.bold:
                font_info["bold"] = True
            if cell.font.italic:
                font_info["italic"] = True
            if cell.font.underline:
                font_info["underline"] = cell.font.underline
            if cell.font.color and cell.font.color.rgb:
                font_info["color"] = cell.font.color.rgb
            if cell.font.size:
                font_info["size"] = cell.font.size
            if font_info:
                formatting["font"] = font_info
        
        # Background color
        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
            formatting["background"] = cell.fill.fgColor.rgb
        
        # Number format
        if cell.number_format and cell.number_format != 'General':
            formatting["number_format"] = cell.number_format
        
        # Alignment
        if cell.alignment:
            align_info = {}
            if cell.alignment.horizontal:
                align_info["horizontal"] = cell.alignment.horizontal
            if cell.alignment.vertical:
                align_info["vertical"] = cell.alignment.vertical
            if cell.alignment.wrap_text:
                align_info["wrap_text"] = True
            if align_info:
                formatting["alignment"] = align_info
        
        return formatting if formatting else None
    
    async def _extract_ods(self,
                          file_path: Path,
                          sheet_names: Optional[List[str]],
                          max_empty_rows: int) -> Dict:
        """Extract content from ODS file."""
        doc = opendocument.load(str(file_path))
        sheets = []
        
        # Get all tables (sheets)
        for table in doc.spreadsheet.getElementsByType(OdfTable):
            table_name = table.getAttribute("name") or "Sheet"
            
            if sheet_names and table_name not in sheet_names:
                continue
            
            sheet_data = await self._extract_ods_sheet(table, max_empty_rows)
            sheet_data["name"] = table_name
            sheets.append(sheet_data)
        
        metadata = {
            "format": "ods",
            "total_sheets": len(sheets),
            "sheet_names": [s["name"] for s in sheets]
        }
        
        return {
            "sheets": sheets,
            "metadata": metadata,
            "statistics": self._calculate_statistics(sheets)
        }
    
    async def _extract_ods_sheet(self, table, max_empty_rows: int) -> Dict:
        """Extract data from ODS table."""
        cells = []
        headers = []
        
        rows = table.getElementsByType(TableRow)
        actual_max_row = 0
        empty_row_count = 0
        
        for row_idx, row in enumerate(rows):
            row_empty = True
            row_cells = []
            
            col_idx = 0
            for cell in row.getElementsByType(TableCell):
                # Handle repeated cells
                repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                
                # Extract cell value
                value = self._extract_ods_cell_value(cell)
                
                if value:
                    row_empty = False
                    empty_row_count = 0
                    
                    for i in range(repeat):
                        cell_data = {
                            "row": row_idx,
                            "col": col_idx + i,
                            "value": value,
                            "type": self._get_ods_cell_type(cell)
                        }
                        
                        # Handle spanning
                        rowspan = int(cell.getAttribute("numberrowsspanned") or 1)
                        colspan = int(cell.getAttribute("numbercolumnsspanned") or 1)
                        if rowspan > 1 or colspan > 1:
                            cell_data["rowspan"] = rowspan
                            cell_data["colspan"] = colspan
                        
                        row_cells.append(cell_data)
                
                col_idx += repeat
            
            if not row_empty:
                cells.extend(row_cells)
                actual_max_row = row_idx
            else:
                empty_row_count += 1
                if empty_row_count >= max_empty_rows:
                    break
        
        # Detect headers
        if cells and row_idx > 0:
            first_row_cells = [c for c in cells if c["row"] == 0]
            if all(c["type"] == "string" for c in first_row_cells):
                headers = [0]
        
        return {
            "rows": actual_max_row + 1,
            "cols": max(c["col"] for c in cells) + 1 if cells else 0,
            "cells": cells,
            "headers": headers
        }
    
    def _extract_ods_cell_value(self, cell) -> str:
        """Extract value from ODS cell."""
        # Try different value types
        value = cell.getAttribute("value")
        if value:
            return value
        
        # Get text content
        text_content = []
        for node in cell.childNodes:
            if node.nodeType == node.TEXT_NODE:
                text_content.append(node.data)
            elif hasattr(node, 'childNodes'):
                for child in node.childNodes:
                    if child.nodeType == child.TEXT_NODE:
                        text_content.append(child.data)
        
        return ' '.join(text_content).strip()
    
    def _get_ods_cell_type(self, cell) -> str:
        """Determine ODS cell type."""
        value_type = cell.getAttribute("valuetype")
        if value_type == "float":
            return "number"
        elif value_type == "date":
            return "date"
        elif value_type == "boolean":
            return "boolean"
        elif value_type == "string":
            return "string"
        else:
            return "string"
    
    def _calculate_statistics(self, sheets: List[Dict]) -> Dict:
        """Calculate extraction statistics."""
        stats = {
            "total_sheets": len(sheets),
            "total_cells": 0,
            "total_rows": 0,
            "total_cols": 0,
            "cell_types": {},
            "sheets_with_charts": 0,
            "sheets_with_merged_cells": 0
        }
        
        for sheet in sheets:
            stats["total_rows"] += sheet.get("rows", 0)
            stats["total_cols"] = max(stats["total_cols"], sheet.get("cols", 0))
            
            cells = sheet.get("cells", [])
            stats["total_cells"] += len(cells)
            
            # Count cell types
            for cell in cells:
                cell_type = cell.get("type", "unknown")
                stats["cell_types"][cell_type] = stats["cell_types"].get(cell_type, 0) + 1
            
            # Count features
            if sheet.get("charts"):
                stats["sheets_with_charts"] += 1
            if sheet.get("merged_cells"):
                stats["sheets_with_merged_cells"] += 1
        
        return stats
    
    def _get_cache_key(self, file_path: Path) -> str:
        """Generate cache key."""
        stat = file_path.stat()
        data = f"{file_path.absolute()}:{stat.st_size}:{stat.st_mtime}"
        return hashlib.sha256(data.encode()).hexdigest()
    
    async def _check_cache(self, cache_key: str) -> Optional[Dict]:
        """Check cache for existing extraction."""
        cache_file = self.cache_dir / f"{cache_key}.json"
        if cache_file.exists():
            try:
                with open(cache_file) as f:
                    return json.load(f)
            except:
                pass
        return None
    
    async def _cache_result(self, cache_key: str, result: Dict):
        """Cache extraction result."""
        cache_file = self.cache_dir / f"{cache_key}.json"
        try:
            with open(cache_file, 'w') as f:
                json.dump(result, f, indent=2)
        except Exception as e:
            logger.warning(f"Failed to cache: {e}")
    
    def display_sheet_preview(self, sheet_data: Dict, max_rows: int = 10, max_cols: int = 10):
        """Display a preview of sheet data as a table."""
        table = RichTable(title=f"Sheet: {sheet_data.get('name', 'Unnamed')}")
        
        # Get cells organized by row/col
        cells = sheet_data.get("cells", [])
        if not cells:
            console.print("[yellow]No data in sheet[/yellow]")
            return
        
        # Organize cells into grid
        grid = {}
        for cell in cells:
            key = (cell["row"], cell["col"])
            grid[key] = cell["value"]
        
        # Determine range to display
        max_display_row = min(max_rows, sheet_data.get("rows", 0))
        max_display_col = min(max_cols, sheet_data.get("cols", 0))
        
        # Add columns
        for col in range(max_display_col):
            col_letter = chr(65 + col) if col < 26 else f"COL{col}"
            table.add_column(col_letter, style="cyan" if col == 0 else None)
        
        # Add rows
        for row in range(max_display_row):
            row_data = []
            for col in range(max_display_col):
                value = grid.get((row, col), "")
                # Truncate long values
                if len(str(value)) > 20:
                    value = str(value)[:17] + "..."
                row_data.append(str(value))
            
            # Highlight header rows
            style = "bold" if row in sheet_data.get("headers", []) else None
            table.add_row(*row_data, style=style)
        
        console.print(table)
        
        # Show summary if truncated
        if sheet_data["rows"] > max_rows or sheet_data["cols"] > max_cols:
            console.print(f"[dim]Showing {max_display_row}x{max_display_col} of "
                        f"{sheet_data['rows']}x{sheet_data['cols']} cells[/dim]")


# Initialize extractor
extractor = SpreadsheetExtractor()


@app.command()
def extract(
    spreadsheet_file: Path = typer.Argument(..., help="Spreadsheet file to extract"),
    output: Optional[Path] = typer.Option(None, "--output", "-o", help="Output JSON file"),
    sheets: Optional[str] = typer.Option(None, "--sheets", "-s", help="Comma-separated sheet names"),
    no_formulas: bool = typer.Option(False, "--no-formulas", help="Don't extract formulas"),
    no_formatting: bool = typer.Option(False, "--no-formatting", help="Don't extract formatting"),
    preview: bool = typer.Option(False, "--preview", "-p", help="Show preview of data")
):
    """Extract content from spreadsheet document."""
    async def run():
        try:
            # Parse sheet names
            sheet_names = None
            if sheets:
                sheet_names = [s.strip() for s in sheets.split(',')]
            
            # Extract
            result = await extractor.extract_spreadsheet(
                spreadsheet_file,
                sheet_names=sheet_names,
                extract_formulas=not no_formulas,
                extract_formatting=not no_formatting
            )
            
            # Show preview if requested
            if preview:
                for sheet in result["sheets"]:
                    extractor.display_sheet_preview(sheet)
                    console.print()
            
            # Save or display summary
            if output:
                with open(output, 'w') as f:
                    json.dump(result, f, indent=2)
                console.print(f"[green] Saved to {output}[/green]")
            else:
                # Display summary
                stats = result["statistics"]
                metadata = result["metadata"]
                
                table = RichTable(title="Spreadsheet Extraction Summary")
                table.add_column("Metric", style="cyan")
                table.add_column("Value", style="green")
                
                table.add_row("Format", metadata["format"].upper())
                table.add_row("Total Sheets", str(stats["total_sheets"]))
                table.add_row("Total Cells", f"{stats['total_cells']:,}")
                table.add_row("Total Rows", str(stats["total_rows"]))
                table.add_row("Total Columns", str(stats["total_cols"]))
                
                # Cell type breakdown
                if stats["cell_types"]:
                    for cell_type, count in stats["cell_types"].items():
                        table.add_row(f"  {cell_type.title()} cells", f"{count:,}")
                
                console.print(table)
                
                # List sheets
                if result["sheets"]:
                    console.print("\n[bold]Sheets:[/bold]")
                    for sheet in result["sheets"]:
                        hidden = " [dim](hidden)[/dim]" if sheet.get("hidden") else ""
                        console.print(f"  " {sheet['name']} ({sheet['rows']}x{sheet['cols']}){hidden}")
        
        except Exception as e:
            console.print(f"[red]Error: {e}[/red]")
            raise typer.Exit(1)
    
    asyncio.run(run())


@app.command()
def analyze(
    spreadsheet_file: Path = typer.Argument(..., help="Spreadsheet to analyze"),
    verbose: bool = typer.Option(False, "--verbose", "-v", help="Show detailed analysis")
):
    """Analyze spreadsheet structure and content."""
    async def run():
        try:
            result = await extractor.extract_spreadsheet(spreadsheet_file)
            
            console.print(f"\n[bold]Analysis of {spreadsheet_file.name}[/bold]\n")
            
            # File info
            metadata = result["metadata"]
            console.print(f"Format: {metadata['format'].upper()}")
            console.print(f"Sheets: {metadata['total_sheets']}")
            
            if metadata.get("creator"):
                console.print(f"Creator: {metadata['creator']}")
            if metadata.get("modified"):
                console.print(f"Modified: {metadata['modified']}")
            
            # Analyze each sheet
            for sheet in result["sheets"]:
                console.print(f"\n[bold cyan]Sheet: {sheet['name']}[/bold cyan]")
                console.print(f"  Dimensions: {sheet['rows']} rows × {sheet['cols']} columns")
                
                cells = sheet.get("cells", [])
                if cells:
                    # Analyze data types
                    type_counts = {}
                    formula_count = 0
                    formatted_count = 0
                    
                    for cell in cells:
                        cell_type = cell.get("type", "unknown")
                        type_counts[cell_type] = type_counts.get(cell_type, 0) + 1
                        
                        if "formula" in cell:
                            formula_count += 1
                        if "formatting" in cell:
                            formatted_count += 1
                    
                    console.print(f"  Cell types: {dict(type_counts)}")
                    if formula_count:
                        console.print(f"  Formulas: {formula_count}")
                    if formatted_count:
                        console.print(f"  Formatted cells: {formatted_count}")
                
                if sheet.get("merged_cells"):
                    console.print(f"  Merged cells: {len(sheet['merged_cells'])}")
                if sheet.get("charts"):
                    console.print(f"  Charts: {len(sheet['charts'])}")
                
                if verbose and cells:
                    # Sample data
                    console.print("\n  [dim]Sample data:[/dim]")
                    extractor.display_sheet_preview(sheet, max_rows=5, max_cols=5)
            
            # Named ranges
            if result.get("named_ranges"):
                console.print(f"\n[bold]Named Ranges:[/bold]")
                for name, info in result["named_ranges"].items():
                    console.print(f"  {name}: {info['range']}")
        
        except Exception as e:
            console.print(f"[red]Error: {e}[/red]")
            raise typer.Exit(1)
    
    asyncio.run(run())


# Worker functions
async def working_usage():
    """Demonstrate spreadsheet extraction."""
    logger.info("Testing spreadsheet extraction...")
    
    # Create test CSV
    test_csv = """Name,Age,Score,Date
Alice,25,95.5,2024-01-01
Bob,30,87.0,2024-01-02
Charlie,35,92.5,2024-01-03
"Smith, John",40,88.0,2024-01-04"""
    
    test_file = Path("/tmp/test_extract.csv")
    with open(test_file, 'w') as f:
        f.write(test_csv)
    
    # Extract
    result = await extractor.extract_spreadsheet(test_file)
    
    logger.info(f"\nExtraction complete:")
    logger.info(f"  Format: {result['metadata']['format']}")
    logger.info(f"  Sheets: {result['statistics']['total_sheets']}")
    logger.info(f"  Total cells: {result['statistics']['total_cells']}")
    
    # Show preview
    if result["sheets"]:
        logger.info("\nPreview:")
        extractor.display_sheet_preview(result["sheets"][0])


async def debug_function():
    """Test edge cases in spreadsheet extraction."""
    logger.info("Testing spreadsheet edge cases...")
    
    if OPENPYXL_AVAILABLE:
        # Create test Excel file
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Test Sheet"
        
        # Add headers
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["C1"] = "Formula"
        
        # Add data with formulas
        ws["A2"] = "Item 1"
        ws["B2"] = 100
        ws["C2"] = "=B2*2"
        
        ws["A3"] = "Item 2"
        ws["B3"] = 200
        ws["C3"] = "=B3*2"
        
        ws["A4"] = "Total"
        ws["B4"] = "=SUM(B2:B3)"
        ws["C4"] = "=SUM(C2:C3)"
        
        # Add formatting
        from openpyxl.styles import Font, PatternFill
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws["C1"].font = Font(bold=True)
        ws["A4"].font = Font(bold=True)
        ws["B4"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Merge cells
        ws.merge_cells("A5:C5")
        ws["A5"] = "This is a merged cell"
        
        test_file = Path("/tmp/test_excel.xlsx")
        wb.save(test_file)
        wb.close()
        
        # Extract with formulas
        result = await extractor.extract_spreadsheet(test_file, extract_formulas=True)
        
        logger.info(f"\nExcel extraction complete:")
        logger.info(f"  Sheets: {len(result['sheets'])}")
        
        # Check formulas
        formula_count = 0
        for sheet in result["sheets"]:
            for cell in sheet.get("cells", []):
                if "formula" in cell:
                    formula_count += 1
                    logger.info(f"  Found formula at ({cell['row']},{cell['col']}): {cell['formula']}")
        
        logger.info(f"  Total formulas: {formula_count}")
        
        # Check merged cells
        if result["sheets"][0].get("merged_cells"):
            logger.info(f"  Merged cells: {len(result['sheets'][0]['merged_cells'])}")
    
    else:
        logger.warning("openpyxl not available - skipping Excel tests")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "working_usage":
        asyncio.run(working_usage())
    elif len(sys.argv) > 1 and sys.argv[1] == "debug":
        asyncio.run(debug_function())
    else:
        app()