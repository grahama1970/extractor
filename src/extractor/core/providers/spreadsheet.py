"""
Module: spreadsheet.py
Purpose: Native spreadsheet extraction without PDF conversion

External Dependencies:
- openpyxl: https://openpyxl.readthedocs.io/  (XLSX/XLS)
- odfpy:      https://pypi.org/project/odfpy/  (ODS)

Example usage
-------------
>>> from extractor.core.providers.spreadsheet import SpreadsheetProvider
>>> doc = SpreadsheetProvider().extract_document("data.xlsx")
>>> print(doc.source_type)   # SourceType.SPREADSHEET
>>> print(len(doc.blocks))   # all tables + images + metadata blocks
"""

import hashlib
import base64
import mimetypes
import os
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Union, Tuple

from openpyxl import load_workbook
from openpyxl.cell import Cell as OpxCell
from odf.table import Table as OdfTable, TableRow, TableCell as OdfCell
from odf.opendocument import load as odf_load

from loguru import logger

from extractor.core.schema.unified_document import (
    UnifiedDocument,
    BlockType,
    SourceType,
    BaseBlock,
    TableBlock,
    ImageBlock,
    BlockMetadata,
    DocumentMetadata,
    HierarchyNode,
    TableCell,
)
from extractor.core.providers.fetcher_bridge import ensure_local_source, attach_fetcher_metadata
from extractor.core.providers.utils import ensure_hierarchy


class SpreadsheetProvider:
    """Direct spreadsheet extraction without PDF conversion."""

    def __init__(self, config: Optional[Dict[str, Any]] = None):
        self.config = config or {}
        self.block_counter = 0

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def extract_document(self, filepath: Union[str, Path]) -> UnifiedDocument:
        resolved_path, fetch_download = ensure_local_source(filepath)
        filepath = Path(resolved_path)
        logger.info(f"Extracting spreadsheet: {filepath}")

        suffix = filepath.suffix.lower()

        if suffix in {".xlsx", ".xlsm", ".xls"}:
            # Default: Use semantic extraction for better document structure
            # XLSX_SIMPLE_MODE: One table per sheet, no semantic parsing
            # XLSX_RAW_MODE: Original openpyxl extraction without semantic parsing
            mode = os.environ.get("XLSX_MODE", "semantic").lower()
            if mode == "simple" or os.environ.get("XLSX_SIMPLE_MODE", "").lower() in {
                "1",
                "true",
                "yes",
            }:
                blocks, metadata = self._extract_openpyxl_simple(filepath)
            elif mode == "raw":
                blocks, metadata = self._extract_openpyxl(filepath)
            else:
                # Default: semantic extraction for better document structure
                blocks, metadata = self._extract_semantic_blocks(filepath)
        elif suffix == ".ods":
            blocks, metadata = self._extract_ods(filepath)
        elif suffix in {".csv", ".tsv"}:
            blocks, metadata = self._extract_csv(filepath, delimiter="\t" if suffix == ".tsv" else ",")
        else:
            raise ValueError(f"Unsupported spreadsheet format: {suffix}")

        attach_fetcher_metadata(metadata, fetch_download)

        # Build hierarchy: Workbook → Worksheets → Tables
        hierarchy = self._build_hierarchy(blocks)

        doc = UnifiedDocument(
            id=self._generate_doc_id(filepath),
            source_type=SourceType.SPREADSHEET,
            source_path=str(filepath),
            blocks=blocks,
            hierarchy=hierarchy,
            metadata=metadata,
            full_text=self._extract_full_text(blocks),
            keywords=[],
        )

        logger.info(f"Extracted {len(blocks)} blocks from spreadsheet")
        return ensure_hierarchy(doc, default_title=filepath.stem)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    def _generate_doc_id(self, filepath: Path) -> str:
        return hashlib.md5(str(filepath).encode()).hexdigest()

    def _generate_block_id(self) -> str:
        self.block_counter += 1
        return f"xls-block-{self.block_counter}"

    # ------------------------------------------------------------------
    # Semantic Structure Detection
    # ------------------------------------------------------------------
    def _is_heading_cell(self, text: str) -> Tuple[bool, int]:
        """Detect if a cell contains a heading and return (is_heading, level).

        Handles:
        - Numbered headings: "4. Architecture", "4.1 Core", "4.1.5.4. BHT"
        - Indented headings: "  4.1 Core" (2 spaces = level 2)
        - Document title row: "Document Title" label
        """
        if not text:
            return False, 0

        text = text.strip()

        # Numbered heading pattern: "4.", "4.1", "4.1.5", "4.1.5.4."
        numbered = re.match(r"^(\d+\.(?:\d+\.?)*)(.+)$", text)
        if numbered:
            num_part = numbered.group(1)
            level = num_part.count(".") if num_part.endswith(".") else num_part.count(".") + 1
            return True, min(level, 6)

        return False, 0

    def _is_requirement_text(self, text: str) -> Optional[str]:
        """Extract requirement ID if text contains a requirement pattern."""
        if not text:
            return None
        match = re.search(r"(REQ-[\w-]+)", text, re.IGNORECASE)
        return match.group(1).upper() if match else None

    def _extract_semantic_blocks(self, filepath: Path) -> Tuple[List[BaseBlock], DocumentMetadata]:
        """Extract semantic structure from XLSX: headings, paragraphs, tables, requirements."""
        wb = load_workbook(filepath, data_only=True, keep_links=False)
        blocks: List[BaseBlock] = []
        doc_title = filepath.stem

        for ws in wb.worksheets:
            sheet_name = ws.title.lower()

            # Overview/Content sheets: parse for document structure
            if any(kw in sheet_name for kw in ["overview", "content", "summary", "toc", "index"]):
                blocks.extend(self._parse_overview_sheet(ws, doc_title))
            # Table sheets: extract as tables
            elif any(kw in sheet_name for kw in ["table", "data", "signal", "param"]):
                table_block = self._openpyxl_sheet_to_table(ws)
                if table_block:
                    blocks.append(table_block)
            else:
                # Default: try semantic parsing first, fall back to table
                semantic_blocks = self._parse_overview_sheet(ws, doc_title)
                if semantic_blocks:
                    blocks.extend(semantic_blocks)
                else:
                    table_block = self._openpyxl_sheet_to_table(ws)
                    if table_block:
                        blocks.append(table_block)

        metadata = DocumentMetadata(
            title=doc_title,
            format_metadata={
                "file_type": "xlsx",
                "creator": wb.properties.creator or "",
                "created": wb.properties.created,
                "modified": wb.properties.modified,
                "sheet_names": wb.sheetnames,
                "file_size": filepath.stat().st_size,
                "semantic_extraction": True,
            },
        )
        return blocks, metadata

    def _parse_overview_sheet(self, ws, doc_title: str) -> List[BaseBlock]:
        """Parse an overview/content sheet extracting headings, paragraphs, requirements."""
        blocks: List[BaseBlock] = []
        current_paragraph_lines: List[str] = []

        def flush_paragraph():
            nonlocal current_paragraph_lines
            if current_paragraph_lines:
                text = "\n".join(current_paragraph_lines)
                req_id = self._is_requirement_text(text)
                blocks.append(
                    BaseBlock(
                        id=self._generate_block_id(),
                        type=BlockType.PARAGRAPH,
                        content=text,
                        metadata=BlockMetadata(
                            attributes=(
                                {"sheet": ws.title, "req_id": req_id}
                                if req_id
                                else {"sheet": ws.title}
                            ),
                            confidence=1.0,
                        ),
                    )
                )
                current_paragraph_lines = []

        for row in ws.iter_rows(values_only=True):
            # Get first non-empty cell content
            text = None
            for cell_val in row:
                if cell_val is not None and str(cell_val).strip():
                    text = str(cell_val).strip()
                    break

            if not text:
                flush_paragraph()
                continue

            # Check for document title row
            if row[0] == "Document Title" and len(row) > 1 and row[1]:
                doc_title = str(row[1])
                blocks.append(
                    BaseBlock(
                        id=self._generate_block_id(),
                        type=BlockType.HEADING,
                        content=doc_title,
                        metadata=BlockMetadata(
                            attributes={"level": 1, "sheet": ws.title}, confidence=1.0
                        ),
                    )
                )
                continue

            # Check for heading
            is_heading, level = self._is_heading_cell(text)
            if is_heading:
                flush_paragraph()
                blocks.append(
                    BaseBlock(
                        id=self._generate_block_id(),
                        type=BlockType.HEADING,
                        content=text,
                        metadata=BlockMetadata(
                            attributes={"level": level, "sheet": ws.title}, confidence=1.0
                        ),
                    )
                )
                continue

            # Regular content - accumulate
            current_paragraph_lines.append(text)

        flush_paragraph()
        return blocks

    # ------------------------------------------------------------------
    # Excel (openpyxl)
    # ------------------------------------------------------------------
    def _extract_openpyxl(self, filepath: Path) -> tuple[List[BaseBlock], DocumentMetadata]:
        wb = load_workbook(filepath, data_only=False, keep_links=False)
        blocks: List[BaseBlock] = []

        for ws in wb.worksheets:
            table_block = self._openpyxl_sheet_to_table(ws)
            if table_block:
                blocks.append(table_block)

        # Images
        for ws in wb.worksheets:
            try:
                images = getattr(ws, "_images", [])
            except Exception:
                images = []
            for img in images:
                img_bytes = img.ref
                mime, _ = mimetypes.guess_type(img.path)
                b64 = base64.b64encode(img_bytes).decode()
                data_uri = f"data:{mime or 'image/png'};base64,{b64}"
                blocks.append(
                    ImageBlock(
                        id=self._generate_block_id(),
                        type=BlockType.IMAGE,
                        content="",
                        src=data_uri,
                        alt=img.path,
                        metadata=BlockMetadata(
                            attributes={"sheet": ws.title, "embedded": True}, confidence=1.0
                        ),
                    )
                )

        metadata = DocumentMetadata(
            format_metadata={
                "file_type": "xlsx",
                "creator": wb.properties.creator or "",
                "created": wb.properties.created,
                "modified": wb.properties.modified,
                "sheet_names": wb.sheetnames,
                "file_size": filepath.stat().st_size,
            }
        )
        return blocks, metadata

    def _openpyxl_sheet_to_table(self, ws) -> Optional[TableBlock]:
        cells: List[TableCell] = []
        headers: List[int] = []

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row == 0 or max_col == 0:
            return None

        for r_idx, row in enumerate(ws.iter_rows(), start=0):
            for c_idx, cell in enumerate(row, start=0):
                value = self._cell_value(cell)
                rowspan, colspan = 1, 1
                if cell.coordinate in ws.merged_cells:
                    # compute merged bounds
                    rng = next(r for r in ws.merged_cells.ranges if cell.coordinate in r)
                    min_row, min_col, max_row_rng, max_col_rng = rng.bounds
                    rowspan = max_row_rng - min_row + 1
                    colspan = max_col_rng - min_col + 1
                cells.append(
                    TableCell(
                        row=r_idx,
                        col=c_idx,
                        content=value,
                        rowspan=rowspan,
                        colspan=colspan,
                    )
                )
            if r_idx == 0:
                headers.append(r_idx)

        return TableBlock(
            id=self._generate_block_id(),
            type=BlockType.TABLE,
            content={},
            rows=max_row,
            cols=max_col,
            cells=cells,
            headers=headers,
            metadata=BlockMetadata(
                attributes={"sheet": ws.title, "source": "openpyxl"}, confidence=1.0
            ),
        )

    def _extract_openpyxl_simple(self, filepath: Path) -> tuple[List[BaseBlock], DocumentMetadata]:
        """Structure-first simple extractor: one table per sheet.

        This keeps XLSX parity informational (not enforced) and avoids exploding
        rows into many paragraph blocks. Suitable for smoke/reporting only.
        """
        wb = load_workbook(filepath, data_only=True, keep_links=False)
        blocks: List[BaseBlock] = []

        for ws in wb.worksheets:
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            cells: List[TableCell] = []
            for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=0):
                for c_idx, val in enumerate(row, start=0):
                    cells.append(
                        TableCell(
                            row=r_idx,
                            col=c_idx,
                            content=str(val) if val not in (None, "") else "",
                            rowspan=1,
                            colspan=1,
                        )
                    )
            blocks.append(
                TableBlock(
                    id=self._generate_block_id(),
                    type=BlockType.TABLE,
                    content="table",
                    rows=max_row,
                    cols=max_col,
                    cells=cells,
                    headers=[0] if max_row > 0 else [],
                    metadata=BlockMetadata(
                        attributes={"sheet": ws.title, "source": "openpyxl_simple"}
                    ),
                )
            )

        metadata = DocumentMetadata(
            format_metadata={
                "file_type": "xlsx",
                "sheet_names": wb.sheetnames,
                "file_size": filepath.stat().st_size,
                "simple": True,
            }
        )
        return blocks, metadata

    def _cell_value(self, cell: OpxCell) -> str:
        return str(cell.value) if cell.value is not None else ""

    # ------------------------------------------------------------------
    # ODS (odfpy)
    # ------------------------------------------------------------------
    def _extract_ods(self, filepath: Path) -> tuple[List[BaseBlock], DocumentMetadata]:
        doc = odf_load(str(filepath))
        blocks: List[BaseBlock] = []

        for table in doc.spreadsheet.getElementsByType(OdfTable):
            table_block = self._ods_sheet_to_table(table)
            if table_block:
                blocks.append(table_block)

        metadata = DocumentMetadata(
            format_metadata={
                "file_type": "ods",
                "sheets": len(doc.spreadsheet.getElementsByType(OdfTable)),
                "file_size": filepath.stat().st_size,
            }
        )
        return blocks, metadata

    def _ods_sheet_to_table(self, table: OdfTable) -> Optional[TableBlock]:
        rows = table.getElementsByType(TableRow)
        cells: List[TableCell] = []
        headers: List[int] = []

        for r_idx, row in enumerate(rows):
            ods_cells = row.getElementsByType(OdfCell)
            for c_idx, cell in enumerate(ods_cells):
                value = str(cell) if cell else ""
                rowspan = int(cell.getAttribute("numberrowsspanned") or 1)
                colspan = int(cell.getAttribute("numbercolumnsspanned") or 1)
                cells.append(
                    TableCell(
                        row=r_idx,
                        col=c_idx,
                        content=value,
                        rowspan=rowspan,
                        colspan=colspan,
                    )
                )
            if r_idx == 0:
                headers.append(r_idx)

        max_row = len(rows)
        max_col = max((c.col + c.colspan - 1 for c in cells), default=0) + 1
        if max_row == 0 or max_col == 0:
            return None

        return TableBlock(
            id=self._generate_block_id(),
            type=BlockType.TABLE,
            content={},
            rows=max_row,
            cols=max_col,
            cells=cells,
            headers=headers,
            metadata=BlockMetadata(
                attributes={"sheet": table.getAttribute("name") or "Sheet"}, confidence=1.0
            ),
        )

    # ------------------------------------------------------------------
    # CSV (native csv)
    # ------------------------------------------------------------------
    def _extract_csv(self, filepath: Path, delimiter: str = ",") -> tuple[List[BaseBlock], DocumentMetadata]:
        import csv
        blocks: List[BaseBlock] = []
        
        # Safety: Check file size. If > 200MB, warn and maybe cap?
        # But 'streaming' means we iterate. The issue is UnifiedDocument stores ALL blocks in memory.
        # We can at least avoid the intermediate 'rows' list double-allocation.
        file_size = filepath.stat().st_size
        if file_size > 200 * 1024 * 1024:
            logger.warning(f"Large CSV detected ({file_size / 1024 / 1024:.1f} MB). memory usage will be high.")

        cells: List[TableCell] = []
        max_col = 0
        row_count = 0

        def _process_reader(reader_obj):
            nonlocal max_col, row_count
            for r_idx, row in enumerate(reader_obj):
                row_count += 1
                if not row: continue
                # Update max_col
                if len(row) > max_col:
                    max_col = len(row)
                
                # Create TableCells immediately, do not store raw row in list
                for c_idx, val in enumerate(row):
                    # Only store non-empty cells to save some memory? 
                    # UnifiedDocument schema expects dense matrix or sparse?
                    # TableCell has row/col, so sparse is fine.
                    # Text usually strips.
                    val_str = str(val).strip()
                    if val_str:
                        cells.append(
                            TableCell(
                                row=r_idx,
                                col=c_idx,
                                content=val_str,
                                rowspan=1,
                                colspan=1,
                            )
                        )

        try:
            # Try sniffing format with a small sample
            with filepath.open(newline='', encoding='utf-8') as f:
                sample = f.read(2048)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                    has_header = csv.Sniffer().has_header(sample)
                except Exception:
                    dialect = None
                
                reader = csv.reader(f, dialect) if dialect else csv.reader(f, delimiter=delimiter)
                _process_reader(reader)

        except Exception:
            # Fallback (different encoding, etc)
            logger.warning("CSV utf-8 read failed, trying latin-1")
            cells = [] # Reset
            row_count = 0
            max_col = 0
            try:
                 with filepath.open(newline='', encoding='latin-1') as f:
                    reader = csv.reader(f)
                    _process_reader(reader)
            except Exception as e:
                logger.error(f"Failed to read CSV: {e}")
                return [], DocumentMetadata(format_metadata={"file_type": "csv", "error": str(e)})

        if not cells and row_count == 0:
             return [], DocumentMetadata(format_metadata={"file_type": "csv", "rows": 0})

        # Create the TableBlock
        # Note: We still create one giant TableBlock. 
        # For true streaming support, UnifiedDocument would need to accept an iterator or we'd need multiple blocks.
        # But this removes the intermediate `rows` list copy, saving ~50% memory.
        table_block = TableBlock(
            id=self._generate_block_id(),
            type=BlockType.TABLE,
            content={},
            rows=row_count,
            cols=max_col,
            cells=cells,
            headers=[0], 
            metadata=BlockMetadata(
                attributes={"sheet": "CSV", "source": "csv_module", "streaming": True}, confidence=1.0
            ),
        )
        blocks.append(table_block)

        metadata = DocumentMetadata(
            format_metadata={
                "file_type": "csv",
                "rows": row_count,
                "file_size": file_size,
            }
        )
        return blocks, metadata

    # ------------------------------------------------------------------
    # Hierarchy
    # ------------------------------------------------------------------
    def _build_hierarchy(self, blocks: List[BaseBlock]) -> Optional[HierarchyNode]:
        root = HierarchyNode(id="root", title="Workbook", level=0, block_id="root", children=[])

        # Group by sheet
        sheets: Dict[str, List[BaseBlock]] = {}
        for b in blocks:
            # Add null check for metadata.attributes
            if b.metadata and b.metadata.attributes:
                sheet = b.metadata.attributes.get("sheet", "Sheet1")
            else:
                sheet = "Sheet1"
            sheets.setdefault(sheet, []).append(b)

        for sheet_name, sheet_blocks in sheets.items():
            sheet_node = HierarchyNode(
                id=f"sheet-{sheet_name}",
                title=sheet_name,
                level=1,
                block_id=f"sheet-{sheet_name}",
                parent_id="root",
                breadcrumb=["Workbook", sheet_name],
            )
            root.children.append(sheet_node)

        return root if root.children else None

    def _extract_full_text(self, blocks: List[BaseBlock]) -> str:
        return "\n".join(
            c.content
            for b in blocks
            if b.type == BlockType.TABLE and hasattr(b, "cells")
            for c in b.cells
        )


# ----------------------------------------------------------------------
# Quick self-test
# ----------------------------------------------------------------------
if __name__ == "__main__":
    import tempfile
    import os

    # Create minimal XLSX
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Demo"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Alice"
    ws["B2"] = 42

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        tmp = f.name

    doc = SpreadsheetProvider().extract_document(tmp)
    assert doc.source_type == SourceType.SPREADSHEET
    assert len(doc.blocks) >= 1
    os.unlink(tmp)
    print("✅ Spreadsheet native provider self-test passed")
