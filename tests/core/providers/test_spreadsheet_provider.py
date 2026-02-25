"""Tests for SpreadsheetProvider with semantic extraction."""

from pathlib import Path

import pytest
from openpyxl import Workbook

from extractor.core.providers.spreadsheet import SpreadsheetProvider
from extractor.core.schema.unified_document import BlockType, SourceType


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    """Create a simple XLSX with basic data."""
    xlsx_path = tmp_path / "simple.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "Item1"
    ws["B2"] = 100
    wb.save(xlsx_path)
    return xlsx_path


@pytest.fixture
def semantic_xlsx(tmp_path: Path) -> Path:
    """Create an XLSX with document-like semantic structure."""
    xlsx_path = tmp_path / "semantic.xlsx"
    wb = Workbook()

    # Overview sheet with document structure
    overview = wb.active
    overview.title = "Overview"
    overview["A1"] = "Document Title"
    overview["B1"] = "Test Document"
    overview["A3"] = "1. Introduction"
    overview["A4"] = "This is the introduction paragraph."
    overview["A5"] = "1.1 Scope"
    overview["A6"] = "The scope includes REQ-001 requirements."
    overview["A7"] = "2. Details"
    overview["A8"] = "More details here."

    # Table sheet
    table_sheet = wb.create_sheet("Table_1")
    table_sheet["A1"] = "Signal"
    table_sheet["B1"] = "Type"
    table_sheet["A2"] = "clk"
    table_sheet["B2"] = "logic"
    table_sheet["A3"] = "rst"
    table_sheet["B3"] = "logic"

    wb.save(xlsx_path)
    return xlsx_path


def test_spreadsheet_provider_basic(simple_xlsx: Path):
    """Test basic spreadsheet extraction."""
    provider = SpreadsheetProvider()
    doc = provider.extract_document(simple_xlsx)

    assert doc.source_type == SourceType.SPREADSHEET
    assert len(doc.blocks) >= 1


def test_spreadsheet_provider_semantic_extraction(semantic_xlsx: Path):
    """Test semantic extraction with headings and paragraphs."""
    provider = SpreadsheetProvider()
    doc = provider.extract_document(semantic_xlsx)

    assert doc.source_type == SourceType.SPREADSHEET

    headings = [b for b in doc.blocks if b.type == BlockType.HEADING]
    paragraphs = [b for b in doc.blocks if b.type == BlockType.PARAGRAPH]
    tables = [b for b in doc.blocks if b.type == BlockType.TABLE]

    # Should extract document title and numbered headings
    assert len(headings) >= 3, f"Expected 3+ headings, got {len(headings)}"

    # Check document title was extracted
    title_blocks = [h for h in headings if "Test Document" in h.content]
    assert len(title_blocks) >= 1, "Document title not extracted"

    # Check numbered headings
    numbered = [h for h in headings if h.content.startswith(("1.", "2."))]
    assert len(numbered) >= 2, "Numbered headings not extracted"

    # Should extract paragraphs
    assert len(paragraphs) >= 1

    # Should extract the table from Table_1 sheet
    assert len(tables) >= 1


def test_spreadsheet_provider_requirement_detection(semantic_xlsx: Path):
    """Test that requirements (REQ-xxx) are detected in paragraphs."""
    provider = SpreadsheetProvider()
    doc = provider.extract_document(semantic_xlsx)

    paragraphs = [b for b in doc.blocks if b.type == BlockType.PARAGRAPH]

    # Find paragraph with requirement
    req_paragraphs = [
        p
        for p in paragraphs
        if p.metadata and p.metadata.attributes and p.metadata.attributes.get("req_id")
    ]

    assert len(req_paragraphs) >= 1, "Requirement ID not detected"
    assert req_paragraphs[0].metadata.attributes["req_id"] == "REQ-001"


def test_spreadsheet_provider_hierarchy(simple_xlsx: Path):
    """Test that hierarchy is built correctly."""
    provider = SpreadsheetProvider()
    doc = provider.extract_document(simple_xlsx)

    assert doc.hierarchy is not None
    assert doc.hierarchy.title == "Workbook" or doc.hierarchy.title == simple_xlsx.stem


def test_spreadsheet_provider_ods(tmp_path: Path):
    """Test ODS extraction (if odfpy is available)."""
    pytest.importorskip("odf")

    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    ods_path = tmp_path / "test.ods"
    doc = OpenDocumentSpreadsheet()

    table = Table(name="Sheet1")
    for row_data in [["A", "B"], ["1", "2"]]:
        tr = TableRow()
        for cell_data in row_data:
            tc = TableCell()
            tc.addElement(P(text=cell_data))
            tr.addElement(tc)
        table.addElement(tr)

    doc.spreadsheet.addElement(table)
    doc.save(ods_path)

    provider = SpreadsheetProvider()
    result = provider.extract_document(ods_path)

    assert result.source_type == SourceType.SPREADSHEET
    assert len(result.blocks) >= 1


def test_spreadsheet_provider_heading_level_detection():
    """Test heading level detection for numbered patterns."""
    provider = SpreadsheetProvider()

    # Test various heading patterns
    test_cases = [
        ("1. Overview", True, 1),
        ("1.1 Scope", True, 2),
        ("1.1.1 Details", True, 3),
        ("4.1.5.4. BHT", True, 4),
        ("Regular paragraph text", False, 0),
        ("REQ-001: Requirement", False, 0),
    ]

    for text, expected_is_heading, expected_level in test_cases:
        is_heading, level = provider._is_heading_cell(text)
        assert is_heading == expected_is_heading, f"'{text}' heading detection failed"
        if expected_is_heading:
            assert (
                level == expected_level
            ), f"'{text}' level detection failed: got {level}, expected {expected_level}"
